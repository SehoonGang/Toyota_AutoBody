import datetime
import json
import math
import os
import shutil
import sys
from typing import Optional, Sequence, Union
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (    
    QApplication, QComboBox, QHBoxLayout,
    QLabel, QMainWindow, QMessageBox, QCheckBox,
    QPushButton, QRadioButton, QTextEdit,
    QVBoxLayout, QWidget, QButtonGroup,
    QLabel, QLineEdit, QPushButton, QTextEdit, QMessageBox
)
from PyQt6.QtOpenGLWidgets import QOpenGLWidget
import pyqtgraph.opengl as gl
import torch
from core import Utils, FileType, PCD
import open3d as o3d
import numpy as np
from tqdm import tqdm
import imageio.v2 as iio
import cv2
from ultralytics import YOLO
import re
import copy
from pathlib import Path

import matplotlib.pyplot as plt
import cv2
import pickle
from datetime import datetime
import pyqtgraph as pg

import math
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dataclasses import dataclass
from typing import Iterable, Optional, Sequence, Union
import socket

from copick3d_api_python import Camera, DeviceState, logging


FORCE_REFRESH = False  # True로 설정하면 캐시 무시하고 재계산

@dataclass
class RoiRow:
    roi_id: Union[int, str]
    cad_xyz: Sequence[float]          # (x,y,z)
    source_xyz: Sequence[float]       # (x,y,z)
    real_ng: bool                     # 실제 NG 여부(너의 로직 결과)
    distance_threshold: Optional[float] = None  # 거리 NG 판정 기준 (None이면 판정 안 함)

class PointCloudView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.view = gl.GLViewWidget()
        self.view.setCameraPosition(distance=2000)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.view)

        self.scatter = None

        g = gl.GLGridItem()
        g.scale(200, 200, 1)
        self.view.addItem(g)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._cad_scale = 1.0,        
        self.setWindowTitle("Body Hole Auto Insepction System")
        self.resize(1920, 1020)
        root = QWidget()
        self.setCentralWidget(root)        
        self.utils = Utils()
        self.pcd = PCD()
        self._set_path()

        self.result_pcd = o3d.geometry.PointCloud()
        self.result_T = np.eye(4, dtype=np.float64)
        self.T_list = []
        self.frame_idx = {}
        self.devices = {}
        self.camera = None
        self.curr_sensor = str
        self.save_sensor_data_path = "./scan"
        self.scan_count = 1
        self.isConnected = False

        self.pose_base_dir_path = "./data/poses"
        self.without_z_base_dir_path = "./data/mask"
        
        leftWidget = QWidget()
        rightWidget = QWidget()        
        layout = QVBoxLayout(root)
        contentLayout = QHBoxLayout()

        contentLayout.addWidget(leftWidget, 3)
        contentLayout.addWidget(rightWidget, 1)
        layout.addLayout(contentLayout)

        leftLayout = QVBoxLayout(leftWidget)
        leftLayout.setContentsMargins(0, 0, 0, 0)
        self.view3d = PointCloudView()
        leftLayout.addWidget(self.view3d)

        rightLayout = QVBoxLayout(rightWidget)
        rightLayout.setAlignment(Qt.AlignmentFlag.AlignTop)

        lblSetting = QLabel("GENERAL")
        lblSetting.setStyleSheet("font-size: 18px; font-weight: bold;")
        rightLayout.addWidget(lblSetting)
        self.radioGroup = QButtonGroup(self)
        # self.radioL = QRadioButton("L Model")
        # self.radioR = QRadioButton("R Model")
        # self.radioR.setChecked(True)
        
        radioRow = QHBoxLayout()
        lblRadioButton = QLabel("MODEL")
        self.radioGroup.addButton(self.radioL)
        self.radioGroup.addButton(self.radioR) 
        radioRow.addWidget(lblRadioButton)
        radioRow.addStretch(1)
        radioRow.addWidget(self.radioL)
        radioRow.addWidget(self.radioR)
        rightLayout.addLayout(radioRow)
        
        calibrationFileRow = QHBoxLayout()        
        self.tbCalibrationFilePath = QLineEdit(self._calib_path)        
        self.btnCalibrationFilePath = QPushButton("LOAD")

        calibrationFileRow.addWidget(QLabel("CALIBRATION"))        
        calibrationFileRow.addWidget(self.tbCalibrationFilePath)
        calibrationFileRow.addWidget(self.btnCalibrationFilePath)
        rightLayout.addLayout(calibrationFileRow)

        deepLearningFileRow = QHBoxLayout()
        deepLearningFileRow.addWidget(QLabel("DEEP LEARNING"))
        # self.tbDeepLearningModelFilePath = QLineEdit(rf"C:\Users\SehoonKang\Desktop\dataset\260113_Scan\260113_Scan\260120_seg_v2.pt")        
        self.tbDeepLearningModelFilePath = QLineEdit(self._seg_model_path)   
        deepLearningFileRow.addWidget(self.tbDeepLearningModelFilePath)
        self.btnDeepLearningFilePath = QPushButton("LOAD")
        deepLearningFileRow.addWidget(self.btnDeepLearningFilePath)
        rightLayout.addLayout(deepLearningFileRow)

        lblScan = QLabel("SCAN")
        lblScan.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px")
        rightLayout.addWidget(lblScan)

        discoverSensorRow = QHBoxLayout()
        self.cmbSensors = QComboBox()        
        self.btnDiscover = QPushButton("DISCOVER")
        self.btnDiscover.setFixedWidth(80)
        discoverSensorRow.addWidget(self.cmbSensors)
        discoverSensorRow.addWidget(self.btnDiscover)
        rightLayout.addLayout(discoverSensorRow)

        connectRow = QHBoxLayout()
        self.btnConnect = QPushButton("CONNECT")
        connectRow.addWidget(self.btnConnect)
        rightLayout.addLayout(connectRow)
        
        scanRow = QHBoxLayout()
        self.btnScan = QPushButton("SCAN [1]")
        self.btnScan.setFixedWidth(380)
        self.btnScan.clicked.connect(self.on_scan)
        self.cbSave = QCheckBox("SAVE")
        self.cbSave.setChecked(True)       
        scanRow.addWidget(self.btnScan)
        scanRow.addWidget(self.cbSave)
        rightLayout.addLayout(scanRow)
        self.btnIcpMerge = QPushButton("MERGE")
        self.btnScanInspect = QPushButton("INSPECT")
        rightLayout.addWidget(self.btnIcpMerge)
        rightLayout.addWidget(self.btnScanInspect)
        self.btnRefresh = QPushButton("REFRESH")        
        rightLayout.addWidget(self.btnRefresh)  

        lblTest = QLabel("TEST")
        lblTest.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px")
        rightLayout.addWidget(lblTest)

        sourceDataFolderRow = QHBoxLayout()
        sourceDataFolderRow.addWidget(QLabel("SOURCE"))
        # self.tbSourceDataFolderPath = QLineEdit(rf"C:\Users\SehoonKang\Desktop\dataset\260113_Scan\260113_Scan\RH")
        self.tbSourceDataFolderPath = QLineEdit(self._source_dir)
        sourceDataFolderRow.addWidget(self.tbSourceDataFolderPath)
        self.btnSourceDataLoad = QPushButton("LOAD")
        sourceDataFolderRow.addWidget(self.btnSourceDataLoad)
        rightLayout.addLayout(sourceDataFolderRow)
        
        self.btnMerge = QPushButton("MERGE")
        self.btnInspect = QPushButton("INSPECT")
        rightLayout.addWidget(self.btnMerge)        
        rightLayout.addWidget(self.btnInspect)        
              
        lblLog = QLabel("LOG")
        lblLog.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 20px")
        rightLayout.addWidget(lblLog)        
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        rightLayout.addWidget(self.log, 1)

        self.cmbSensors.currentIndexChanged.connect(self.on_select_sensor)
        self.btnDiscover.clicked.connect(self.on_discover_sensors)
        self.btnSourceDataLoad.clicked.connect(self.on_source_data_load)
        self.btnCalibrationFilePath.clicked.connect(self.on_calibration_file_load)
        self.btnDeepLearningFilePath.clicked.connect(self.on_deep_learning_file_load)        
        self.btnMerge.clicked.connect(self.on_merge)
        self.btnInspect.clicked.connect(self.on_inspect)
        self.btnRefresh.clicked.connect(self.on_refresh)
        self.btnIcpMerge.clicked.connect(self.on_icp_merge_and_move_to_cad)
        self.btnScanInspect.clicked.connect(self.on_scan_inspect)
        self.btnConnect.clicked.connect(self.on_connect_sensor)

        self.utils.on_load_source_data_folder(self.tbCalibrationFilePath.text(), FileType.Calibration)
        self.log.append(rf"Loaded Calibration file({self.tbCalibrationFilePath.text()}) succesfully.")

        self.utils.on_load_source_data_folder(self.tbDeepLearningModelFilePath.text(), FileType.DeepLearningModel)
        self.log.append(rf"Loaded Deep Learning Model File({self.tbDeepLearningModelFilePath.text()}) successfully.")        
        self.seg_model = YOLO(self.tbDeepLearningModelFilePath.text())

    def _set_path(self):
        config = load_cfg()
        # os 맞게 다시한번 확인
        self._source_dir = str(Path(config["source_data_folder"]))
        self._calib_path = str(Path(config["calibration_file"]))
        self._seg_model_path = str(Path(config["deep_learning_model"]))

        self.pcd.set_path(body_path=str(Path(config["bodyPath"])))
        body_pose = str(config["BodyPosition"])

        self.radioL = QRadioButton("L MODEL")
        self.radioR = QRadioButton("R MODEL")
        if body_pose == "Left":
            self.radioL.setChecked(True)
        elif body_pose == "Right":
            self.radioR.setChecked(True)
        else:
            print(f"Body Pose UnKnown Left or Right yout Input: {body_pose}")
            print("set Right Pose default")
            self.radioR.setChecked(True)
        self._cad_scale = float(config["cad_scale"]),   
    
    def on_icp_merge_and_move_to_cad(self):
        T_list, merged_pcd, reference_pcd = self.pcd.scan_icp_merge_pcd()

        self.T_list = T_list
        cad_centers_array = np.array(self.utils.cad_data[self.current_model()]["cad_centers"], dtype=np.float32)
        moved_merge_pcd, T_to_cad, report = self.pcd.move_merged_pcd_to_cad(merged_pcd=merged_pcd,
                                                                            CAD_CENTERS=cad_centers_array,
                                                                            align_points=np.asarray(reference_pcd, dtype=np.float64),
                                                                            copy_pcd=True)
        
        self.result_pcd = moved_merge_pcd
        self.result_T = T_to_cad
        self.log.append("Merged and moved the merged pcd to cad successfully.")
    
    def on_discover_sensors(self):
        self.devices = {}
        
        for device in Camera.discover_devices() :
            serial_number = device.serial_number
            state = device.query_device_state()
            self.log.append(rf"{serial_number} / {state}")
            
            if state == DeviceState.Ready :
                self.devices[device.serial_number] = device

            self.cmbSensors.clear()
            self.cmbSensors.addItems(self.devices.keys())
            self.curr_sensor = self.devices[list(self.devices.keys())[0]]

    def on_connect_sensor(self) :
        if self.isConnected :
            self.camera.disconnect()
            self.camera = None
            self.isConnected = False
            self.btnConnect.setText(rf"CONNECT")
        else :
            if self.curr_sensor.query_device_state() == DeviceState.Ready :
                self.camera = Camera()
                self.camera.connect(self.curr_sensor)
                self.log.append(rf"Connect Camera : Serial({self.curr_sensor.serial_number})")
                self.isConnected = True
                self.btnConnect.setText(rf"DISCONNECT")
        self.scan_count  = 1
        self.btnScan.setText(rf"SCAN [{self.scan_count}]")
    
    def on_select_sensor(self):        
        self.curr_sensor = self.devices[self.cmbSensors.currentText()]
        print(">>>>>>>>>>>>>>>> " , self.curr_sensor)

    def on_source_data_load(self):        
        self.utils.on_load_source_data_folder(self.tbSourceDataFolderPath.text(), FileType.Image)
        self.log.append(rf"load {self.tbSourceDataFolderPath.text()} completed.")

    def on_calibration_file_load(self):
        self.utils.on_load_source_data_folder(self.tbCalibrationFilePath.text(), FileType.Calibration)
        self.log.append(rf"load {self.tbCalibrationFilePath.text()} completed.")

    def on_deep_learning_file_load(self):
        self.utils.on_load_source_data_folder(self.tbDeepLearningModelFilePath.text(), FileType.DeepLearningModel)
        self.log.append(rf"load {self.tbDeepLearningModelFilePath.text()} completed.")
        #변경 필요 ========================================
        self.seg_model = YOLO(self.tbDeepLearningModelFilePath.text())

    def on_refresh(self):
        self.scan_count  = 1
        self.btnScan.setText(rf"SCAN [{self.scan_count}]")

    def on_scan(self):
        self.tbScanCount.setText(str(self.scan_count))        
        frame = self.camera.scan_frame()

        texture = frame.get_texture()
        point_map = frame.get_point_map()
        x_point_map = point_map[:, :, 0].astype(np.float32)
        y_point_map = point_map[:, :, 1].astype(np.float32)
        z_point_map = point_map[:, :, 2].astype(np.float32)
        pose_path = Path(rf"{self.pose_base_dir_path}/{self.current_model()}/{self.scan_count}_SCAN_POSE.txt")

        without_z_file_path = rf"{self.without_z_base_dir_path}/{self.current_model()}/{self.scan_count}_IMG_PointCloud_Z.tif"
        path = rf"{self.save_sensor_data_path}/{self.current_model()}/"#/{datetime.now().strftime("%Y%m%d_%H%M%S")}"
        mask_save_path = rf"{path}/{self.scan_count}_Mask.tiff"
        mask_array = self.utils.create_mask_from_depth_array(z_point_map, without_z_file_path, mask_save_path = mask_save_path, tolerance=10, min_area = 10, use_morph=True, save = self.cbSave.isChecked())

        self.pcd.scan_merge_pcd(self.utils.calibration_file_path, "fanuc",  self.current_model(),
                                frame_number=self.scan_count, texture=texture,
                                point_x=x_point_map, point_y=y_point_map, point_z=z_point_map,
                                mask= mask_array, pose_path=pose_path)
        self.log.append(rf"[Frame {self.scan_count}] Merged pcd for successfully")        

        # normal_map = frame.get_normal_map()
        # x_normal_map = normal_map[:, :, 0].astype(np.float32)
        # cv2.imwrite(rf"{path}/{self.scan_count}_IMG_NormalMap_X.tif", x_normal_map)
        # y_normal_map = normal_map[:, :, 1].astype(np.float32)
        # cv2.imwrite(rf"{path}/{self.scan_count}_IMG_NormalMap_Y.tif", y_normal_map)
        # z_normal_map = normal_map[:, :, 2].astype(np.float32)
        # cv2.imwrite(rf"{path}/{self.scan_count}_IMG_NormalMap_Z.tif", z_normal_map)
        # self.log.append(rf"Saved {self.scan_count}_NormalMap files files successfully")

        if self.cbSave.isChecked() :
            if not os.path.exists(path):
                os.makedirs(path)
                print(f"폴더가 생성되었습니다: {path}")

            cv2.imwrite(rf"{path}/{self.scan_count}_IMG_Texture_8Bit.png", texture)
            dst_dir = Path(rf"{path}")
            dst_dir.mkdir(parents=True, exist_ok = True)
            shutil.copy2(str(pose_path), str(dst_dir/pose_path.name))            

            cv2.imwrite(rf"{path}/{self.scan_count}_IMG_PointCloud_X.tif", x_point_map)
            cv2.imwrite(rf"{path}/{self.scan_count}_IMG_PointCloud_Y.tif", y_point_map)
            cv2.imwrite(rf"{path}/{self.scan_count}_IMG_PointCloud_Z.tif", z_point_map)
            self.log.append(rf"[Frame [{self.scan_count}]] Saved scan data successfully")

        self.scan_count += 1
        self.btnScan.setText(rf"SCAN [{self.scan_count}]")

    def on_merge(self):
        self.log.append("[INFO] Start to merge frames")

        # load Cache Start
        cache_file = f"cache/merge_pcd_{self.current_model()}.pkl"

        cache_dir = "cache"
        os.makedirs(cache_dir, exist_ok=True)
        cache_file = f"{cache_dir}/merge_pcd_{self.current_model()}.pkl"
        merged_pcd_file = f"{cache_dir}/merged_pcd_{self.current_model()}.ply"
        
        if not FORCE_REFRESH and os.path.exists(cache_file):
            print("Loading from cache...")
            with open(cache_file, 'rb') as f:
                cached_data = pickle.load(f)
            T_list = cached_data['T_list']
            reference_pcd = cached_data['reference_pcd']
            merged_pcd = o3d.io.read_point_cloud(merged_pcd_file)
        else:
            print("No Cache Runnign merge_pcd func")
            T_list, merged_pcd, reference_pcd = self.pcd.merge_pcd(
                self.utils.source_data_folder_files,
                self.utils.calibration_file_path,
                "fanuc", 
                self.current_model()
            )
            
            # 캐시 저장
            print("Saving to cache...")
            o3d.io.write_point_cloud(merged_pcd_file, merged_pcd)
            with open(cache_file, 'wb') as f:
                pickle.dump({
                    'T_list': T_list,
                    'reference_pcd': reference_pcd
                }, f)                

        self.T_list = T_list
        cad_centers_array = np.array(self.utils.cad_data[self.current_model()]["cad_centers"], dtype=np.float32)
        moved_merge_pcd, T_to_cad, report = self.pcd.move_merged_pcd_to_cad(merged_pcd=merged_pcd,
                                                                            CAD_CENTERS=cad_centers_array,
                                                                            align_points=np.asarray(reference_pcd, dtype=np.float64),
                                                                            copy_pcd=True)
        
        self.result_pcd = moved_merge_pcd
        self.result_T = T_to_cad

        # pcd_base = copy.deepcopy(self.result_pcd)
        cad_points = np.array(self.utils.cad_data[self.current_model()]["cad_welding_points"], dtype=np.float32)
        
        pcd_cad = o3d.geometry.PointCloud()
        pcd_cad.points = o3d.utility.Vector3dVector(cad_points.astype(np.float64))
        pcd_cad.paint_uniform_color((1.0, 0.0, 0.0))

        # self.set_pointcloud(pcd_input=[self.result_pcd, pcd_cad], size=0.5, sampling_rate=0.3)
        self.set_inspected_result_pointcloud(pcd_input=self.result_pcd, center_input=pcd_cad, size= 0.5, sampling_rate=0.3)
        self.log.append("merge frames complete.")
        
    def on_scan_inspect(self) :
        self.log.append("Inspecting data...")
        roi_hole_points_dict = {}
        frame_pcd = {}
        pose_dict = {}
        self.frame_idx = {}

        sorted_scans = sorted(self.pcd.scan_path_dict.items())

        for i, frame in enumerate(tqdm(sorted_scans, total=len(sorted_scans))):
            pcd = PCD()
            texture, point_x, point_y, point_z, pose_path, mask_array = frame

            pts_cam = pcd.scan_make_cam_pcd(point_x=point_x, point_y=point_y, point_z=point_z, texture=texture, mask_array=mask_array)
            frame_number = i+1
            X = point_x
            Y = point_y
            Z = point_z

            mask_valid = np.isfinite(X) & np.isfinite(Y) & np.isfinite(Z)
            mask_valid = np.asarray(mask_valid, dtype=bool)
            mask_nonzero = (X != 0) | (Y != 0) | (Z != 0)
            mask_nonzero = np.asarray(mask_nonzero, dtype=bool)
            mask_valid &= mask_nonzero

            ys_idx, xs_idx = np.where(mask_valid)
            pts_cam = np.stack([
                X[ys_idx, xs_idx],
                Y[ys_idx, xs_idx],
                Z[ys_idx, xs_idx]
            ], axis=1)

            image_bgr = texture            
            bgr = image_bgr[ys_idx, xs_idx]
            rgb = bgr[:, ::-1].astype(np.float64) / 255.0 
            
            T_cam_to_world = self.T_list[i]
            T_world_to_cad = self.result_T
            T_cam_to_cad = T_world_to_cad @ T_cam_to_world
            self.frame_idx[frame_number] = i

            pts_cam = self.transform_points(pts_cam, T_cam_to_world)   
            pts_cam = self.transform_points(pts_cam, T_world_to_cad)

            frame_pcd[frame_number] = {
                "points" : pts_cam,
                "rgb" : rgb,
                "ys_idx" : ys_idx,
                "xs_idx" : xs_idx
            }

            image_for_seg = texture
            img_h, img_w, _ = image_for_seg.shape

            cad_points  = np.array(self.utils.cad_data[self.current_model()]["cad_welding_points"], dtype=np.float32)            
            pcd_cad = o3d.geometry.PointCloud()
            pcd_cad.points = o3d.utility.Vector3dVector(cad_points.astype(np.float64))
            
            
            pp = np.asarray(pts_cam, dtype=np.float64).reshape(-1, 3)
            pcd_pts_cad = o3d.geometry.PointCloud()
            pcd_pts_cad.points = o3d.utility.Vector3dVector(pp)

            seg_pad = 40

            for roi_id, center in enumerate(cad_points, start=1):
                dist = np.linalg.norm(pts_cam - center, axis=1)                
                mask_roi_3d = dist <= 4
                num_roi_pts = np.count_nonzero(mask_roi_3d)

                if num_roi_pts == 0:
                    continue

                roi_y = ys_idx[mask_roi_3d]
                roi_x = xs_idx[mask_roi_3d]

                y_min, y_max = int(roi_y.min()), int(roi_y.max())
                x_min, x_max = int(roi_x.min()), int(roi_x.max())

                pad = seg_pad
                y_min = max(y_min - pad, 0)
                x_min = max(x_min - pad, 0)
                y_max = min(y_max + pad, img_h - 1)
                x_max = min(x_max + pad, img_w - 1)

                if y_max <= y_min or x_max <= x_min:
                    print(rf"{frame_number} // {roi_id} : ymx_{y_max} / ymn_{y_min} / xmx_{x_max} / xmn_{x_min}")
                    continue

                crop_img = image_for_seg[y_min:y_max + 1, x_min:x_max + 1]

                if crop_img.size == 0:                    
                    continue
                
                ch, cw, _ = crop_img.shape
                if ch < 16 or cw < 16:                    
                    continue


                results = self.seg_model(crop_img, device='cuda:0', verbose=False)
                # cv2.imwrite(rf"C:\Users\SehoonKang\Desktop\s\RH\crop_{frame_number}_{roi_id}.png", crop_img)
                # for i, result in enumerate(results):
                #     res_img = result.plot()
                #     cv2.imwrite(rf"C:\Users\SehoonKang\Desktop\s\RH\seg_{frame_number}_{roi_id}.png", res_img)

                if len(results) == 0 or results[0].masks is None:
                    continue

                masks_yolo = results[0].masks.data.cpu().numpy()
                if masks_yolo.shape[0] == 0:
                    print(f"[INFO] ROI : mask 개수 0 (view {i}).")
                    continue

                mask_bin = (masks_yolo > 0.5)
                full_mask_local = np.any(mask_bin, axis=0)
                Hm, Wm = full_mask_local.shape

                if (Hm, Wm) != (ch, cw):
                    full_mask_local = cv2.resize(full_mask_local.astype(np.uint8), (cw, ch), interpolation=cv2.INTER_NEAREST).astype(bool)                 
                
                mask_resized = full_mask_local
                in_crop_2d = (
                    (ys_idx >= y_min) & (ys_idx <= y_max) &
                    (xs_idx >= x_min) & (xs_idx <= x_max)
                )

                if not np.any(in_crop_2d):
                    continue

                ys_c = ys_idx[in_crop_2d] - y_min
                xs_c = xs_idx[in_crop_2d] - x_min
                mask_on_pixels_small = mask_resized[ys_c, xs_c]


                pcd_pts_np = np.asarray(pcd_pts_cad.points, dtype=np.float64)  # (N,3)

                # (B) seg mask로 1차 3D 포인트 추출 (여기서 중심을 구함)
                seg_pts = pcd_pts_np[in_crop_2d][mask_on_pixels_small]
                if seg_pts.shape[0] < 10:
                    # seg는 있는데 3D로 매핑되는 점이 너무 적으면 fallback
                    continue

                center3d = seg_pts.mean(axis=0)

                # (C) 중심 기준 3D in_crop 정의 (반경 r3d는 튜닝)
                r3d = 40.0  # mm 기준 (너 데이터에 맞게 20~80 사이로 시작)
                dist3d = np.linalg.norm(pcd_pts_np - center3d[None, :], axis=1)
                in_crop3d = dist3d <= r3d

                # (D) 최종 포인트: seg 마스크 & 3D 중심 ROI
                #     여기서도 2D seg mask가 필요하니까, 2D seg 조건을 전체 프레임으로 확장해야 함
                #     가장 쉬운 방법: in_crop2d에서 seg mask True였던 인덱스들을 "전역 인덱스"로 복원
                global_idx_in_crop2d = np.where(in_crop_2d)[0]
                global_idx_seg = global_idx_in_crop2d[np.where(mask_on_pixels_small)[0]]

                # seg에 해당하는 전역 bool 마스크 생성 (N,)
                seg_mask_global = np.zeros(pcd_pts_np.shape[0], dtype=bool)
                seg_mask_global[global_idx_seg] = True

                # 최종
                final_mask = seg_mask_global & in_crop3d  # (N,)
                roi_hole_pts = pcd_pts_np[final_mask]
                n_hole = roi_hole_pts.shape[0]
                
                if n_hole > 0:
                    roi_hole_points_dict.setdefault(roi_id, {})[frame_number] = roi_hole_pts

            pose = [float(x) for x in re.findall(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', open(pose_path, 'r', encoding='utf-8').read())]
            pose = pose[:6]
            pose_dict[frame_number] = pose

        self.inspect_real_welding_point(roi_hole_points_dict=roi_hole_points_dict, frame_pcd=frame_pcd, pad=5)

    def on_inspect(self):
        self.log.append("Inspecting data...")
        roi_hole_points_dict = {}        

        def extract_index(group):
            fname = os.path.basename(group[0])
            m = re.match(r"(\d+)_", fname)
            return int(m.group(1)) if m else -1

        source_data_folder_files_sort = sorted(self.utils.source_data_folder_files, key=extract_index)

        frame_pcd = {}
        pose_dict = {}
        self.frame_idx = {} 

        for i, frame in enumerate(tqdm(source_data_folder_files_sort, total=len(source_data_folder_files_sort))):
            pcd = PCD()
            texture_path, x_path, y_path, z_path, pose_path, mask_path = frame

            pts_cam = pcd._make_cam_pcd(x_path=x_path, y_path=y_path, z_path=z_path,texture_path=texture_path, mask_path=mask_path)
            frame_number = os.path.basename(texture_path).replace("_IMG_Texture_8Bit.png", "")
            X = np.asarray(iio.imread(x_path).astype(np.float64))
            Y = np.asarray(iio.imread(y_path).astype(np.float64))
            Z = np.asarray(iio.imread(z_path).astype(np.float64))

            mask_valid = np.isfinite(X) & np.isfinite(Y) & np.isfinite(Z)
            mask_valid = np.asarray(mask_valid, dtype=bool)
            mask_nonzero = (X != 0) | (Y != 0) | (Z != 0)
            mask_nonzero = np.asarray(mask_nonzero, dtype=bool)
            mask_valid &= mask_nonzero

            ys_idx, xs_idx = np.where(mask_valid)
            pts_cam = np.stack([
                X[ys_idx, xs_idx],
                Y[ys_idx, xs_idx],
                Z[ys_idx, xs_idx]
            ], axis=1)

            image_bgr = cv2.imread(texture_path, cv2.IMREAD_COLOR)
            if image_bgr is None:
                raise ValueError(f"Failed to read texture: {texture_path}")
            
            bgr = image_bgr[ys_idx, xs_idx]
            rgb = bgr[:, ::-1].astype(np.float64) / 255.0 
            
            T_cam_to_world = self.T_list[i]
            T_world_to_cad = self.result_T
            T_cam_to_cad = T_world_to_cad @ T_cam_to_world
            self.frame_idx[frame_number] = i

            pts_cam = self.transform_points(pts_cam, T_cam_to_world)   
            pts_cam = self.transform_points(pts_cam, T_world_to_cad)

            frame_pcd[frame_number] = {
                "points" : pts_cam,
                "rgb" : rgb,
                "ys_idx" : ys_idx,
                "xs_idx" : xs_idx
            }

            image_for_seg = cv2.imread(texture_path, cv2.IMREAD_COLOR)
            img_h, img_w, _ = image_for_seg.shape

            cad_points  = np.array(self.utils.cad_data[self.current_model()]["cad_welding_points"], dtype=np.float32)            
            pcd_cad = o3d.geometry.PointCloud()
            pcd_cad.points = o3d.utility.Vector3dVector(cad_points.astype(np.float64))
            
            
            pp = np.asarray(pts_cam, dtype=np.float64).reshape(-1, 3)
            pcd_pts_cad = o3d.geometry.PointCloud()
            pcd_pts_cad.points = o3d.utility.Vector3dVector(pp)

            seg_pad = 40

            for roi_id, center in enumerate(cad_points, start=1):
                dist = np.linalg.norm(pts_cam - center, axis=1)                
                mask_roi_3d = dist <= 4
                num_roi_pts = np.count_nonzero(mask_roi_3d)

                if num_roi_pts == 0:
                    continue

                roi_y = ys_idx[mask_roi_3d]
                roi_x = xs_idx[mask_roi_3d]

                y_min, y_max = int(roi_y.min()), int(roi_y.max())
                x_min, x_max = int(roi_x.min()), int(roi_x.max())

                pad = seg_pad
                y_min = max(y_min - pad, 0)
                x_min = max(x_min - pad, 0)
                y_max = min(y_max + pad, img_h - 1)
                x_max = min(x_max + pad, img_w - 1)

                if y_max <= y_min or x_max <= x_min:
                    print(rf"{frame_number} // {roi_id} : ymx_{y_max} / ymn_{y_min} / xmx_{x_max} / xmn_{x_min}")
                    continue

                crop_img = image_for_seg[y_min:y_max + 1, x_min:x_max + 1]

                if crop_img.size == 0:                    
                    continue
                
                ch, cw, _ = crop_img.shape
                if ch < 16 or cw < 16:                    
                    continue


                results = self.seg_model(crop_img, device='cuda:0', verbose=False)
                # cv2.imwrite(rf"C:\Users\SehoonKang\Desktop\s\RH\crop_{frame_number}_{roi_id}.png", crop_img)
                # for i, result in enumerate(results):
                #     res_img = result.plot()
                #     cv2.imwrite(rf"C:\Users\SehoonKang\Desktop\s\RH\seg_{frame_number}_{roi_id}.png", res_img)

                if len(results) == 0 or results[0].masks is None:
                    continue

                masks_yolo = results[0].masks.data.cpu().numpy()
                if masks_yolo.shape[0] == 0:
                    print(f"[INFO] ROI : mask 개수 0 (view {i}).")
                    continue

                mask_bin = (masks_yolo > 0.5)
                full_mask_local = np.any(mask_bin, axis=0)
                Hm, Wm = full_mask_local.shape

                if (Hm, Wm) != (ch, cw):
                    full_mask_local = cv2.resize(full_mask_local.astype(np.uint8), (cw, ch), interpolation=cv2.INTER_NEAREST).astype(bool)                 
                
                mask_resized = full_mask_local
                in_crop_2d = (
                    (ys_idx >= y_min) & (ys_idx <= y_max) &
                    (xs_idx >= x_min) & (xs_idx <= x_max)
                )

                if not np.any(in_crop_2d):
                    continue

                ys_c = ys_idx[in_crop_2d] - y_min
                xs_c = xs_idx[in_crop_2d] - x_min
                mask_on_pixels_small = mask_resized[ys_c, xs_c]


                pcd_pts_np = np.asarray(pcd_pts_cad.points, dtype=np.float64)  # (N,3)

                # (B) seg mask로 1차 3D 포인트 추출 (여기서 중심을 구함)
                seg_pts = pcd_pts_np[in_crop_2d][mask_on_pixels_small]
                if seg_pts.shape[0] < 10:
                    # seg는 있는데 3D로 매핑되는 점이 너무 적으면 fallback
                    continue

                center3d = seg_pts.mean(axis=0)

                # (C) 중심 기준 3D in_crop 정의 (반경 r3d는 튜닝)
                r3d = 40.0  # mm 기준 (너 데이터에 맞게 20~80 사이로 시작)
                dist3d = np.linalg.norm(pcd_pts_np - center3d[None, :], axis=1)
                in_crop3d = dist3d <= r3d

                # (D) 최종 포인트: seg 마스크 & 3D 중심 ROI
                #     여기서도 2D seg mask가 필요하니까, 2D seg 조건을 전체 프레임으로 확장해야 함
                #     가장 쉬운 방법: in_crop2d에서 seg mask True였던 인덱스들을 "전역 인덱스"로 복원
                global_idx_in_crop2d = np.where(in_crop_2d)[0]
                global_idx_seg = global_idx_in_crop2d[np.where(mask_on_pixels_small)[0]]

                # seg에 해당하는 전역 bool 마스크 생성 (N,)
                seg_mask_global = np.zeros(pcd_pts_np.shape[0], dtype=bool)
                seg_mask_global[global_idx_seg] = True

                # 최종
                final_mask = seg_mask_global & in_crop3d  # (N,)
                roi_hole_pts = pcd_pts_np[final_mask]
                n_hole = roi_hole_pts.shape[0]
                
                if n_hole > 0:
                    roi_hole_points_dict.setdefault(roi_id, {})[frame_number] = roi_hole_pts

            pose = [float(x) for x in re.findall(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', open(pose_path, 'r', encoding='utf-8').read())]
            pose = pose[:6]
            pose_dict[frame_number] = pose

        self.inspect_real_welding_point(roi_hole_points_dict=roi_hole_points_dict, frame_pcd=frame_pcd, pad=5)
        # pcd_base = copy.deepcopy(self.result_pcd)        
        # pcd_base = self.result_pcd.voxel_down_sample(0.5)

        # pcd_holes = self.roi_dict_to_pcd(roi_hole_points_dict=roi_hole_points_dict)

        # vis = o3d.visualization.Visualizer()
        # vis.create_window()
        # vis.add_geometry(pcd_base)
        # vis.add_geometry(pcd_holes)        
        # opt = vis.get_render_option()
        # opt.point_size = 4.0
        # vis.run()
        # vis.destroy_window()

        # self.set_pointcloud(pcd_holes)

    def set_pointcloud(self, pcd_input, *, size: float = 5.0, sampling_rate : float = 0.5):
        if isinstance(pcd_input, list):            
            combined_pcd = o3d.geometry.PointCloud()
            for p in pcd_input:
                combined_pcd += p
            pcd = combined_pcd
        else:
            pcd = pcd_input

        pcd = pcd.voxel_down_sample(sampling_rate)
        cl, ind = pcd.remove_statistical_outlier(nb_neighbors=20, std_ratio=2.0)
        pcd = pcd.select_by_index(ind)
        pts = np.asarray(pcd.points, dtype=np.float32)

        if len(pts) == 0:
            return

        if pcd.has_colors():
            cols = np.asarray(pcd.colors, dtype=np.float32)
            if cols.max() > 1.0:
                cols = cols / 255.0
            alpha = np.ones((cols.shape[0], 1), dtype=np.float32) * 0.4
            cols = np.concatenate([cols, alpha], axis=1)
        else:
            cols = np.ones((pts.shape[0], 4), dtype=np.float32) * 0.5
            cols[:, 3] = 0.4

        if getattr(self.view3d, "scatter", None) is not None:
            self.view3d.view.removeItem(self.view3d.scatter)
            self.view3d.scatter = None
        
        self.view3d.scatter = gl.GLScatterPlotItem(pos=pts, color=cols, size=float(size), pxMode=False)
        self.view3d.view.addItem(self.view3d.scatter)

        center = np.mean(pts, axis=0)
        self.view3d.view.opts['center'] = pg.Vector(center[0], center[1], center[2])
        dist = np.linalg.norm(pts.max(axis=0) - pts.min(axis=0))
        self.view3d.view.setCameraPosition(distance=dist * 0.5)

    def set_inspected_result_pointcloud(self, pcd_input, center_input, *, size: float = 0.5, sampling_rate : float = 0.3):
        if isinstance(pcd_input, list):            
            combined_pcd = o3d.geometry.PointCloud()
            for p in pcd_input:
                combined_pcd += p
            pcd = combined_pcd
        else:
            pcd = pcd_input

        pcd = pcd.voxel_down_sample(sampling_rate)
        cl, ind = pcd.remove_statistical_outlier(nb_neighbors=20, std_ratio=2.0)
        pcd = pcd.select_by_index(ind)
        pts = np.asarray(pcd.points, dtype=np.float32)

        if len(pts) == 0:
            return

        if pcd.has_colors():
            cols = np.asarray(pcd.colors, dtype=np.float32)
            if cols.max() > 1.0:
                cols = cols / 255.0
            alpha = np.ones((cols.shape[0], 1), dtype=np.float32) * 0.4
            cols = np.concatenate([cols, alpha], axis=1)
        else:
            cols = np.ones((pts.shape[0], 4), dtype=np.float32) * 0.5
            cols[:, 3] = 1.0

        if getattr(self.view3d, "scatter", None) is not None:
            self.view3d.view.removeItem(self.view3d.scatter)
            self.view3d.scatter = None
        
        self.view3d.scatter = gl.GLScatterPlotItem(pos=pts, color=cols, size=float(size), pxMode=False)
        self.view3d.view.addItem(self.view3d.scatter)

        if getattr(self.view3d, "gizmo", None) is not None:
            self.view3d.view.removeItem(self.view3d.gizmo)

        if center_input is not None:
            # 1. 기존 Gizmo 그룹 제거
            if getattr(self.view3d, "gizmo_group", None) is not None:
                for item in self.view3d.gizmo_group:
                    self.view3d.view.removeItem(item)
            self.view3d.gizmo_group = []
            c_pts = np.asarray(center_input.points)
            
            if len(c_pts) > 0:
                data_range = np.linalg.norm(pts.max(axis=0) - pts.min(axis=0))
                gizmo_size = data_range * 0.01 # 화살표 크기 조절

                for pt in c_pts:
                    # 50개 포인트 각각에 대해 X(빨강), Y(초록), Z(파랑) 화살표 생성
                    # X축 (Red)
                    x_arrow = CustomGizmo.create_arrow(size=gizmo_size, color=(1, 0, 0, 1), axis='x')
                    x_arrow.translate(pt[0], pt[1], pt[2])
                    
                    # Y축 (Green)
                    y_arrow = CustomGizmo.create_arrow(size=gizmo_size, color=(0, 1, 0, 1), axis='y')
                    y_arrow.translate(pt[0], pt[1], pt[2])
                    
                    # Z축 (Blue)
                    z_arrow = CustomGizmo.create_arrow(size=gizmo_size, color=(0, 0, 1, 1), axis='z')
                    z_arrow.translate(pt[0], pt[1], pt[2])

                    # 화면에 추가 및 리스트 저장
                    for arrow in [x_arrow, y_arrow, z_arrow]:
                        self.view3d.view.addItem(arrow)
                        self.view3d.gizmo_group.append(arrow)

        center = np.mean(pts, axis=0)
        self.view3d.view.opts['center'] = pg.Vector(center[0], center[1], center[2])
        dist = np.linalg.norm(pts.max(axis=0) - pts.min(axis=0))
        self.view3d.view.setCameraPosition(distance=dist * 0.5)
    def roi_dict_to_pcd(self, roi_hole_points_dict: dict[int, dict[int, np.ndarray]]) -> o3d.geometry.PointCloud:
        all_pts = []
        
        # 첫 번째 dict: roi_id, 두 번째 dict: 내부 index 또는 key
        for roi_id, hole_dict in roi_hole_points_dict.items():
            for frame_id, pts in hole_dict.items(): # 리스트 대신 딕셔너리 순회
                if pts is None or len(pts) == 0:
                    print(rf">>>>>>> point count is zero : {roi_id}_{frame_id}")
                    continue
                
                # pts가 이미 (N, 3) 형태의 numpy array라고 가정합니다.
                all_pts.append(np.asarray(pts, dtype=np.float64))

        if not all_pts:
            return o3d.geometry.PointCloud()

        # 모든 포인트를 하나로 합칩니다.
        P = np.vstack(all_pts)

        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(P)
        
        # 가시성을 위해 색상 지정 (연두색)
        pcd.paint_uniform_color([0.0, 1.0, 0.2])
        return pcd    

    def current_model(self):        
        return 'LH' if self.radioL.isChecked() else 'RH'

    def transform_points(self, points_xyz: np.ndarray, T: np.ndarray) -> np.ndarray:
        points_xyz = np.asarray(points_xyz, dtype=np.float64)
        if points_xyz.ndim == 1:
            points_xyz = points_xyz.reshape(1, 3)

        ones = np.ones((points_xyz.shape[0], 1), dtype=np.float64)
        ph = np.hstack([points_xyz, ones])            # (N,4)
        out = (T @ ph.T).T[:, :3]                     # (N,3)
        return out
    
    def inspect_real_welding_point(self,                                 
                                   roi_hole_points_dict : dict[int, dict[int, np.ndarray]],
                                   frame_pcd : dict[int, dict],
                                   pad: int = 30) : 
        samples = []
        src_points = {}
        welding_pcd_dict = {}
        for roi_id, pcd_dict in roi_hole_points_dict.items():
            print(rf"{roi_id}번째 타흔 확인")

            # 1. frame 별로 가장 point cloud 수가 많은 것의
            if not pcd_dict:
                print("  [WARN] pcd_dict empty")
                continue
            
            best_frame_id, best_roi_pts = max(
                pcd_dict.items(),
                key=lambda kv: 0 if kv[1] is None else int(np.asarray(kv[1]).shape[0])
            )

            frame_data = frame_pcd.get(best_frame_id, None)
            if frame_data is None:
                print(f"[ROI {roi_id}] best_frame_pcd is None (frame_id={best_frame_id})")
                continue
            
            frame_pts = np.asarray(frame_data.get("points", None), dtype=np.float64) if frame_data.get("points", None) is not None else None
            frame_rgb = frame_data.get("rgb", None)

            if frame_pts is None or frame_pts.size == 0:
                print(f"[ROI {roi_id}] frame_pts empty (frame_id={best_frame_id})")
                continue
            frame_pts = frame_pts.reshape(-1, 3)

            # rgb가 있으면 shape 맞는지 체크 (없으면 None 유지)
            if frame_rgb is not None:
                frame_rgb = np.asarray(frame_rgb, dtype=np.float64)
                if frame_rgb.shape[0] != frame_pts.shape[0]:
                    print(f"[WARN] rgb length mismatch: pts={frame_pts.shape[0]} rgb={frame_rgb.shape[0]} -> ignore rgb")
                    frame_rgb = None

            # 2. 평면 피팅
            center = np.asarray(best_roi_pts, dtype=np.float64).mean(axis=0)            
            dist = np.linalg.norm(frame_pts - center[None, :], axis=1)
            crop_idx = np.where(dist <= float(pad))[0]

            if crop_idx.size == 0:
                print(f"[ROI {roi_id}] crop empty (pad={pad})")
                continue

            crop_pts = frame_pts[crop_idx]
            pcd_crop = o3d.geometry.PointCloud()
            pcd_crop.points = o3d.utility.Vector3dVector(crop_pts)

            if frame_rgb is not None:
                crop_rgb = frame_rgb[crop_idx]  # (Nc,3)
                pcd_crop.colors = o3d.utility.Vector3dVector(np.clip(crop_rgb, 0.0, 1.0))

            pcd_filtered, inlier_idx = pcd_crop.remove_radius_outlier(
                nb_points=30,
                radius=1.0
            )

            pcd_near, pcd_far, plane, is_welding, w_xyz = self.filter_points_near_plane(pcd_filtered,
                                                                     distance_threshold=0.05,
                                                                     frame_idx = self.frame_idx[best_frame_id],
                                                                     center=center,
                                                                     min_max_threshold = 0.1)
            welding_pcd_dict[roi_id] = pcd_far
            
            if w_xyz is None or np.asarray(w_xyz).size != 3:
                print(f"[ROI {roi_id}] w_xyz invalid: {w_xyz} (is_welding={is_welding})")
                continue

            cad_center_point = self.utils.cad_data[self.current_model()]["cad_welding_points"][roi_id - 1]
            cad_center_point = np.asarray(cad_center_point, dtype=np.float64).reshape(3)
            source_center_point = np.asarray(w_xyz, dtype=np.float64).reshape(3)
            src_points[roi_id] = source_center_point

            distance = self.distance_3d(cad_center_point, source_center_point)

            def sphere_at(p, radius=2.0, color=(1.0, 0.0, 0.0)):
                p = np.asarray(p, dtype=np.float64).reshape(3)
                s = o3d.geometry.TriangleMesh.create_sphere(radius=float(radius))
                s.translate(p)
                s.paint_uniform_color(color)
                return s
            
            w_sphere   = sphere_at(w_xyz, radius=0.5, color=(0.5, 0.5, 0.5))   # 빨강: w_xyz
            cad_sphere = sphere_at(cad_center_point, radius=0.5, color=(0.0, 0, 1)) 

            print(rf"Welding Point {roi_id} : CAD X : {cad_center_point[0]} / CAD Y : {cad_center_point[1]} / CAD Z : {cad_center_point[2]}")
            print(rf"Welding Point {roi_id} : SRC X : {source_center_point[0]} / SRC Y : {source_center_point[1]} / SRC Z : {source_center_point[2]}")
            print(rf"Welding Point {roi_id} : DIST : {distance}")
            print(rf"{roi_id} >>>>> {is_welding}")
            
            # if roi_id == 49 :
            #     pcd_near.paint_uniform_color((1.0, 0.0, 0.0))
            #     o3d.visualization.draw_geometries([pcd_far, pcd_near, w_sphere, cad_sphere],window_name=f"ROI {roi_id}")

            samples.append(RoiRow(roi_id=roi_id,
                                  cad_xyz=(cad_center_point[0], cad_center_point[1], cad_center_point[2]),
                                  source_xyz=(source_center_point[0], source_center_point[1], source_center_point[2]),
                                  real_ng=is_welding,   
                                  distance_threshold= 4))
            
        cad_list = []
        src_list = []
        valid_roi = []

        for roi_id in range(1, 51):
            if roi_id not in src_points:
                continue

            cad_point = np.asarray(np.array(self.utils.cad_data[self.current_model()]["cad_welding_points"], dtype=np.float32)[roi_id-1], dtype=np.float64).reshape(3,)
            src_point = np.asarray(src_points[roi_id], dtype=np.float64).reshape(3,)

            cad_list.append(cad_point)
            src_list.append(src_point)
            valid_roi.append(roi_id)

        def estimate_rigid_kabsch(src, dst):
            src = np.asarray(src, float)
            dst = np.asarray(dst, float)

            src_mean = src.mean(axis=0)
            dst_mean = dst.mean(axis=0)

            X = src - src_mean
            Y = dst - dst_mean

            H = X.T @ Y
            U, S, Vt = np.linalg.svd(H)
            R = Vt.T @ U.T
            if np.linalg.det(R) < 0:
                Vt[-1, :] *= -1
                R = Vt.T @ U.T

            t = dst_mean - R @ src_mean
            return R, t
        
        # cad = np.vstack(cad_list)
        # src = np.vstack(src_list)
        # R, t = estimate_rigid_kabsch(src, cad)
        # src_rigid = (R @ src.T).T + t

        # err_before = np.linalg.norm(src - cad, axis=1)
        # err_after  = np.linalg.norm(src_rigid - cad, axis=1)

        # for i in range(len(src)):
        #     cx, cy, cz = cad[i]
        #     sx, sy, sz = src[i]
        #     px, py, pz = src_rigid[i]

        #     print(f"[{i:02d}] CAD      : {cx:9.3f} {cy:9.3f} {cz:9.3f}")
        #     print(f"     SRC      : {sx:9.3f} {sy:9.3f} {sz:9.3f}   dist={err_before[i]:.3f}")
        #     print(f"     SRC_CORR : {px:9.3f} {py:9.3f} {pz:9.3f}   dist={err_after[i]:.3f}")
        #     print("-"*70)

        samples.sort(key=lambda r: r.roi_id)
        # self.export_roi_distance_excel(samples, rf"{self.current_model()}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        self.export_roi_distance_excel(samples, rf"{self.current_model()}_report.xlsx")        

        geoms = [self.result_pcd]
        # merged_far = o3d.geometry.PointCloud()
        # for roi_id, pcd in welding_pcd_dict.items() :
        #     merged_far += pcd
        # geoms.append(merged_far)
        cad_list = np.array(self.utils.cad_data[self.current_model()]["cad_welding_points"], dtype=np.float32)
        cad_list = cad_list * self._cad_scale 

        src_list = self.register_with_icp(src_list, cad_list)

        for cad in np.array(self.utils.cad_data[self.current_model()]["cad_welding_points"], dtype=np.float32) :
            n_sphere = sphere_at(cad, radius=1, color=(0, 1 ,0))
            geoms.append(n_sphere)

        # for cad in np.array(self.utils.cad_data[self.current_model()]["cad_centers"], dtype=np.float32) :
        #     n_sphere = sphere_at(cad, radius=1, color=(0, 0 ,1))
        #     geoms.append(n_sphere)
        
        for src in src_list :
            sphere = sphere_at(src, radius=1, color=(1, 0 ,0))
            geoms.append(sphere)
        o3d.visualization.draw_geometries(geoms, window_name="ALL ROIs cached")

    def points_to_pcd(self, points_xyz, color=(1.0, 0.0, 0.0)):
        pts = np.asarray(points_xyz, dtype=np.float64).reshape(-1, 3)
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(pts)
        pcd.paint_uniform_color(color)
        return pcd

    def register_with_icp(self, source_list, target_list, threshold=0.05):
        """
        ICP 알고리즘을 사용하여 source를 target에 정합합니다.
        - source_list: 옮길 좌표 리스트 [[x,y,z], ...]
        - target_list: 기준 좌표 리스트 [[x,y,z], ...]
        - threshold: 대응점을 찾을 최대 거리 (이 거리 안의 점들만 정합에 사용)
        """
        # 1. 리스트를 Open3D의 PointCloud 객체로 변환
        source_pcd = o3d.geometry.PointCloud()
        source_pcd.points = o3d.utility.Vector3dVector(np.array(source_list))

        target_pcd = o3d.geometry.PointCloud()
        target_pcd.points = o3d.utility.Vector3dVector(np.array(target_list))

        # 2. 초기 변환 행렬 (Identity Matrix - 변환 없음 상태로 시작)
        trans_init = np.identity(4)

        # 3. ICP 정합 실행 (Point-to-Point 방식)
        print("ICP 정합 중...")
        reg_p2p = o3d.pipelines.registration.registration_icp(
            source_pcd, target_pcd, threshold, trans_init,
            o3d.pipelines.registration.TransformationEstimationPointToPoint()
        )

        # 4. 결과 행렬 확인 (4x4 Transformation Matrix)
        print("최종 변환 행렬:\n", reg_p2p.transformation)

        # 5. 변환 행렬을 source에 적용하여 좌표 이동
        source_pcd.transform(reg_p2p.transformation)

        # 6. 이동된 좌표를 다시 리스트로 변환하여 반환
        registered_points = np.asarray(source_pcd.points).tolist()
        
        return registered_points    

    def filter_points_near_plane(self, pcd, distance_threshold=1.0, ransac_thresh=1.0,
                             ransac_n=3, num_iterations=2000,
                             frame_idx: int | None = None,
                             center: np.ndarray | None = None,
                             trim_ratio: float = 0.05,
                             count_threshold: int = 30,
                             min_max_threshold: float = 0.2):

        pts = np.asarray(pcd.points, dtype=np.float64)
        if pts.size == 0:
            return None, None, np.array([], dtype=np.int64), False, None

        # 1) plane fit
        plane_model, inliers = pcd.segment_plane(
            distance_threshold=float(ransac_thresh),
            ransac_n=int(ransac_n),
            num_iterations=int(num_iterations),
        )
        a, b, c, d = plane_model
        n = np.array([a, b, c], dtype=np.float64)
        n_norm = np.linalg.norm(n)
        if n_norm < 1e-12:
            raise RuntimeError("Plane normal is near zero.")

        # 2) 법선 부호 고정 (cam -> object)
        if frame_idx is not None and center is not None:
            T_cam_to_world = self.T_list[frame_idx]
            T_cam_to_cad = self.result_T @ T_cam_to_world
            cam_pos_cad = T_cam_to_cad[:3, 3]

            center = np.asarray(center, dtype=np.float64).reshape(3)
            v_ref = center - cam_pos_cad  # cam -> object (필요하면 반대로 바꿔 테스트)

            if np.dot(n, v_ref) < 0:
                n = -n
                d = -d
                plane_model = [n[0], n[1], n[2], d]

        # 3) signed distance
        n_norm = (np.linalg.norm(n) + 1e-12)
        signed = (pts @ n + d) / n_norm

        keep = signed > float(distance_threshold)        
        keep_idx = np.where(keep)[0]
        rem_idx  = np.where(~keep)[0]

        def trim_indices_by_signed(global_idx: np.ndarray, signed_all: np.ndarray, trim_ratio: float) -> np.ndarray:            
            if global_idx.size == 0:
                return global_idx

            s = signed_all[global_idx]  # 해당 그룹의 signed 값들
            lo = np.percentile(s, 100.0 * trim_ratio)
            hi = np.percentile(s, 100.0 * (1.0 - trim_ratio))

            keep_local = (s >= lo) & (s <= hi)
            return global_idx[keep_local]

        trim_ratio = float(trim_ratio)  # 파라미터로 받고 있다고 가정
        keep_idx_trim = trim_indices_by_signed(keep_idx, signed, trim_ratio)
        rem_idx_trim  = trim_indices_by_signed(rem_idx,  signed, trim_ratio)

        below = pcd.select_by_index(keep_idx_trim.tolist())
        above = pcd.select_by_index(rem_idx_trim.tolist())

        # 1. above filtering
        above = self.remove_nearest_percent_from_above(above, below, 30)

        # 2. above plane fitting
        plane_model2, n2, above_inlier, above_outlier = self.fit_plane_from_pcd(above, ransac_thresh=0.05)
        a2, b2, c2, d2 = plane_model2
        n2 = np.array([a2, b2, c2], dtype=np.float64)

        if frame_idx is not None and center is not None:
            T_cam_to_world = self.T_list[frame_idx]
            T_cam_to_cad = self.result_T @ T_cam_to_world
            cam_pos_cad = T_cam_to_cad[:3, 3]

            center = np.asarray(center, dtype=np.float64).reshape(3)
            v_ref = center - cam_pos_cad 

            if np.dot(n2, v_ref) < 0:
                n2 = -n2
                d2 = -d2
                plane_model2 = [float(n2[0]), float(n2[1]), float(n2[2]), float(d2)]

        # 3. divide by plane2 
        a2, b2, c2, d2 = plane_model2
        n2 = np.array([a2, b2, c2], dtype=np.float64)
        n2_norm = np.linalg.norm(n2) + 1e-12

        signed2 = (pts @ n2 + d2) / n2_norm
        near_thr2 = 0.03
        near_mask = signed2 > float(near_thr2)
        far_mask  = ~near_mask

        near_idx = np.where(near_mask)[0]
        far_idx  = np.where(far_mask)[0]

        welding_pcd = pcd.select_by_index(near_idx.tolist())
        plane_pcd  = pcd.select_by_index(far_idx.tolist())

        if len(welding_pcd.points) == 0:
            return welding_pcd, plane_pcd, plane_model2, False, None

        # if len(welding_pcd.points) >= 10:
        #     eps = self.auto_eps(welding_pcd, factor=2.5)
        #     welding_pcd = self.keep_largest_spatial_component(welding_pcd, eps, min_points=10)

        if len(welding_pcd.points) == 0:
            return welding_pcd, plane_pcd, plane_model2, False, None

        # welding_pcd.paint_uniform_color([1, 0, 0]) 
        # plane_pcd.paint_uniform_color([0, 1, 0])
        # o3d.visualization.draw_geometries([
        #     welding_pcd,
        #     plane_pcd
        # ])

        # 4. confirm real welding points 
        n_welding = len(welding_pcd.points)
        n_plane   = len(plane_pcd.points)

        if n_welding < count_threshold or n_plane < count_threshold:
            return welding_pcd, plane_pcd, plane_model2, False, None

        # plane1 기준 signed를 welding/plane 각각 다시 계산
        w_pts = np.asarray(welding_pcd.points, dtype=np.float64)
        p_pts = np.asarray(plane_pcd.points, dtype=np.float64)

        s_welding = (w_pts @ n2 + d2) / n2_norm
        s_plane   = (p_pts @ n2 + d2) / n2_norm
        
        weld_ref  = float(np.percentile(s_welding, 90.0))        
        is_welding = weld_ref >= float(min_max_threshold)
        print(rf">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> gap = {weld_ref}")

        # 5. center point
        w_center = w_pts.mean(axis=0)
        w_x, w_y, w_z = float(w_center[0]), float(w_center[1]), float(w_center[2])

        # a2, b2, c2, d2 = plane_model2
        # if abs(c2) < 1e-12:
        #     return welding_pcd, plane_pcd, plane_model2, is_welding, None

        # w_z = -(a2 * w_x + b2 * w_y + d2) / c2
        w_xyz = np.array([w_x, w_y, w_z], dtype=np.float64)

        return welding_pcd, plane_pcd, plane_model2, is_welding, w_xyz
    
    def fit_plane_from_pcd(self, pcd: o3d.geometry.PointCloud,
                       ransac_thresh: float = 0.05,
                       ransac_n: int = 3,
                       num_iterations: int = 2000):
        if pcd is None or len(pcd.points) < ransac_n:
            raise ValueError("Not enough points to fit a plane.")

        plane_model, inliers = pcd.segment_plane(
            distance_threshold=float(ransac_thresh),
            ransac_n=int(ransac_n),
            num_iterations=int(num_iterations),
        )
        a, b, c, d = plane_model
        n = np.array([a, b, c], dtype=np.float64)
        n = n / (np.linalg.norm(n) + 1e-12)

        pcd_inlier = pcd.select_by_index(inliers)
        pcd_outlier = pcd.select_by_index(inliers, invert=True)

        return plane_model, n, pcd_inlier, pcd_outlier    
    
    def remove_nearest_percent_from_above(
        self,
        above: o3d.geometry.PointCloud,
        below: o3d.geometry.PointCloud,
        n_percent: float = 20.0  # 가까운 n% 제거
    ) -> o3d.geometry.PointCloud:
        A = np.asarray(above.points, dtype=np.float64)
        B = np.asarray(below.points, dtype=np.float64)

        if A.size == 0:
            return o3d.geometry.PointCloud()
        if B.size == 0:
            # below가 비어있으면 기준점이 없으니 above 그대로
            return above

        # 1) below 평균점(중심)
        c = B.mean(axis=0)  # (3,)

        # 2) above 각 점의 중심까지 거리
        dist = np.linalg.norm(A - c[None, :], axis=1)

        # 3) 가까운 n% 인덱스 계산
        n_percent = float(n_percent)
        n_percent = max(0.0, min(100.0, n_percent))
        k = int(np.floor(len(dist) * (n_percent / 100.0)))

        if k <= 0:
            return above  # 제거할 게 없음
        if k >= len(dist):
            return o3d.geometry.PointCloud()  # 전부 제거

        order = np.argsort(dist)      # 가까운 순
        remove_idx = order[:k]        # 가까운 k개 제거

        # 4) 제거 후 남길 인덱스
        keep_mask = np.ones(len(dist), dtype=bool)
        keep_mask[remove_idx] = False
        keep_idx = np.where(keep_mask)[0]

        return above.select_by_index(keep_idx.tolist())
    
    def keep_largest_spatial_component(self, pcd: o3d.geometry.PointCloud, eps: float = 2.0, min_points: int = 10):
        
        if pcd is None or len(pcd.points) == 0:
            return o3d.geometry.PointCloud()

        labels = np.asarray(pcd.cluster_dbscan(eps=float(eps), min_points=int(min_points), print_progress=False), dtype=np.int32)

        print("N:", len(labels), "noise:", np.sum(labels == -1), "clusters:", len(set(labels)) - (1 if -1 in labels else 0))

        # 전부 노이즈(-1)면 빈 결과
        valid = labels >= 0
        if not np.any(valid):
            return o3d.geometry.PointCloud()

        counts = np.bincount(labels[valid])
        largest_label = int(np.argmax(counts))

        keep_idx = np.where(labels == largest_label)[0]
        return pcd.select_by_index(keep_idx.tolist())
    
    def auto_eps(self, pcd: o3d.geometry.PointCloud, factor: float = 2.5, sample: int = 5000) -> float:
        if len(pcd.points) == 0:
            return 1.0
        if len(pcd.points) > sample:
            idx = np.random.choice(len(pcd.points), sample, replace=False)
            p = pcd.select_by_index(idx.tolist())
        else:
            p = pcd
        d = np.asarray(p.compute_nearest_neighbor_distance())
        if d.size == 0:
            return 1.0
        return float(np.median(d) * factor)

    def distance_3d(self, p1, p2):
        p1 = np.asarray(p1, dtype=np.float64).reshape(3)
        p2 = np.asarray(p2, dtype=np.float64).reshape(3)
        return float(np.linalg.norm(p1 - p2))

    def export_roi_distance_excel(
        self,
        rows: Iterable[RoiRow],
        out_xlsx_path: str,
        sheet_name: str = "ROI_Report",
    ) -> str:
        """
        ROI별 CAD 좌표, Source 좌표, 거리, NG 판단을 엑셀로 저장.
        distance = Euclidean distance between cad_xyz and source_xyz.
        distance_ng = (distance > distance_threshold) if threshold provided else None

        returns: saved xlsx path
        """
        headers = [
            "Welding ID",
            "Cad x", "Cad y", "Cad z",
            "Real x", "Real y", "Real z",
            "Distance",
            "Real Welding",
            "Distance OK",
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header styling
        header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Data rows
        red_fill = PatternFill("solid", fgColor="FFC7CE")     # light red
        red_font = Font(color="9C0006", bold=True)

        def _dist(a, b) -> float:
            dx = float(a[0]) - float(b[0])
            dy = float(a[1]) - float(b[1])
            dz = float(a[2]) - float(b[2])
            return math.sqrt(dx*dx + dy*dy + dz*dz)

        r = 2
        for item in rows:
            cad = item.cad_xyz
            src = item.source_xyz
            dist = _dist(cad, src)

            if item.distance_threshold is None:
                dist_ng = ""  # 기준 없으면 공란
            else:
                dist_ng = bool(dist < float(item.distance_threshold))

            ws.append([
                str(item.roi_id),
                float(cad[0]), float(cad[1]), float(cad[2]),
                float(src[0]), float(src[1]), float(src[2]),
                float(dist),
                bool(item.real_ng),
                dist_ng if dist_ng != "" else "",
            ])

            # Align + number formats
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).alignment = center

            # 좌표/거리 소수점 포맷
            for c in range(2, 9):  # cad/source/distance
                ws.cell(row=r, column=c).number_format = "0.000"

            # NG 강조(둘 중 하나라도 True면 행을 붉게)
            real_ng_val = bool(item.real_ng)
            dist_ng_val = bool(dist_ng) if dist_ng != "" else False
            if real_ng_val == False or dist_ng_val == False:
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = red_fill
                    # 글씨도 강조하고 싶으면:
                    if c in (9, 10):  # real ng, distance ng
                        cell.font = red_font

            r += 1

        # Column width autosize (simple)
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 28)

        ws.freeze_panes = "A2"
        wb.save(out_xlsx_path)
        return out_xlsx_path

class CustomGizmo:
    @staticmethod
    def create_arrow(size=0.1, color=(1, 0, 0, 1), axis='x'):
        """선이 굵고 끝에 삼각뿔이 달린 화살표 하나를 생성합니다."""
        # 1. 막대(Cylinder) 부분 생성 (이 녀석을 부모로 쓸 겁니다)
        cyl_data = gl.MeshData.cylinder(rows=10, cols=20, radius=[0.05, 0.05], length=1.0)
        cyl = gl.GLMeshItem(meshdata=cyl_data, smooth=True, color=color, shader='shaded', glOptions='opaque')
        
        # 2. 삼각뿔(Cone) 부분 생성
        cone_data = gl.MeshData.cylinder(rows=10, cols=20, radius=[0.1, 0.0], length=0.3)
        cone = gl.GLMeshItem(meshdata=cone_data, smooth=True, color=color, shader='shaded', glOptions='opaque')
        
        # 삼각뿔을 막대 끝(length=1.0)으로 이동시킨 후 막대의 자식으로 등록
        cone.translate(0, 0, 1.0)
        cone.setParentItem(cyl) 
        
        # 3. 축에 따른 회전 설정
        # 기본은 Z축 방향이므로 x, y일 때만 회전시킵니다.
        if axis == 'x':
            cyl.rotate(90, 0, 1, 0)
        elif axis == 'y':
            cyl.rotate(-90, 1, 0, 0)
            
        # 4. 전체 크기 조절
        cyl.scale(size, size, size)
        
        return cyl # 이제 cyl이 화살표 전체를 대표합니다.
def load_cfg():

    config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    
    with open(config_path, 'r', encoding='utf-8') as f:
        all_configs = json.load(f)
    
    # 현재 컴퓨터 이름 가져오기
    hostname = socket.gethostname()

    print(f"Computer Name : {hostname}")
    
    # 해당 컴퓨터의 설정 가져오기 (없으면 default 사용)
    config = all_configs.get(hostname, all_configs.get("default", {}))
    
    print(f"Loading config for: {hostname}")
    return config

def main():
    
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()

    sys.exit(app.exec())

if __name__ == "__main__":    
    main()